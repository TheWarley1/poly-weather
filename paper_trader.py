"""
Auto Paper-Trade Pipeline
==========================
Fully automated: scans all markets, computes edges, logs "paper trades" for
every signal that passes the filters, and after resolution auto-fills outcomes.

Writes results to paper_trade_tracker.xlsx (the same file you can open in Excel).
Also writes a JSON log (paper_trades.json) as the source of truth.

Usage:
  python paper_trader.py              # scan + resolve (run every 6 hours alongside forecast_logger.py)
  python paper_trader.py --status     # print performance summary
  python paper_trader.py --export     # re-export JSON → Excel

Schedule alongside forecast_logger.py:
  python forecast_logger.py && python paper_trader.py
"""

import requests
import csv
import json
import os
import re
import sys
import time
import threading
import numpy as np
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path
from scipy import stats
from scipy.stats import skewnorm
from zoneinfo import ZoneInfo

# Parallelism for the city scan loop. Every per-city API call is I/O-bound
# (Polymarket + Open-Meteo + NWS), so threads give a near-linear speedup.
SCAN_WORKERS = 8
# Lock guarding print() so multi-line city blocks from scan_city_edges
# don't tear across threads.
_scan_print_lock = threading.Lock()

# Import shared infrastructure from forecast_logger
from forecast_logger import (
    CITIES, GAMMA_BASE, MONTH_SLUG, MODEL_WEIGHTS,
    LOG_DIR, NWS_HEADERS,
    load_forecast_log, city_today,
    build_event_slug, fetch_polymarket_event, get_resolution_winner,
    fetch_openmeteo_multimodel_highs,
    compute_ensemble_high,
    is_sane_temp,
    retry_get,  # PATCH L
)

# ─────────────────────────────────────────────
# CONFIG — paper trading parameters
# ─────────────────────────────────────────────
CONFIG = {
    "min_ev": 0.08,           # 8% minimum edge (was 5% — too loose, kept firing on noise)
    "min_price": 0.08,        # NEVER buy below 8¢. Cheap long shots destroy bankrolls.
    "max_price": 0.45,        # Never buy above 45¢ (avoid favorites)
    "min_volume": 1000,       # Minimum market volume ($)
    "kelly_fraction": 0.25,   # Quarter Kelly
    "max_shares": 200,        # Cap share count (prevents 4000-share long-shot disasters)
    "min_hours": 2.0,         # Min hours to resolution
    "max_hours": 72.0,        # Max hours to resolution
    "min_calibration_obs": 5, # Require 5+ paired forecast/actual observations before trading a city
    "min_edge_narrow_bin": 0.12,  # Narrow bins (2°F/1°C wide) need 12% edge, not 8%
    "min_model_prob": 0.03,       # PATCH E: never trade bins where our model says <3% prob.
                                  # Tails are where the model is least trustworthy — any
                                  # "edge" found there is more likely a σ-estimation
                                  # artifact than a real mispricing.

    # ──────────────────────────────────────────────────────────
    # PATCH F: bankroll-realistic sizing.
    #
    # The prior flat-$ model had three flaws we hit in production:
    #   1) `bankroll: 1000.0` never updated — Kelly% was scaled to a
    #      fictitious constant regardless of P&L. A bad week didn't
    #      shrink bet sizes.
    #   2) `max_bet: 15.0` dominated every bet with real edge — 38
    #      scans yesterday all landed at the $15 cap, putting $570
    #      (57% of stated bankroll) at risk in a single session.
    #   3) No daily exposure cap and no drawdown circuit breaker.
    #
    # The replacement: live bankroll = starting + realized P&L; every
    # sizing cap is a FRACTION of that live bankroll; drawdowns from
    # start-of-session bankroll throttle Kelly; a margin-of-safety
    # shrink pulls model prob toward market (discount for our own σ
    # uncertainty).
    # ──────────────────────────────────────────────────────────
    "starting_bankroll": 1000.0,  # Paper starting stake (fixed; live bankroll computed from P&L)
    "max_bet_frac": 0.05,         # Per-trade cap: 5% of LIVE bankroll
    "max_daily_exposure": 0.25,   # Total open-trade cost cannot exceed 25% of live bankroll
    "drawdown_soft": 0.99,        # At 20% drawdown, halve Kelly
    "drawdown_hard": 1,        # At 40% drawdown, stop opening new trades
    "prob_shrinkage": 0.10,       # Pull model_prob 10% toward market price before sizing.
                                  # Margin of safety: we don't trust our σ to the last decimal,
                                  # so we bet a slightly more conservative edge than the raw model says.
}

TRADES_FILE = Path(__file__).parent / "paper_trades.json"
EXCEL_FILE = Path(__file__).parent / "paper_trade_tracker.xlsx"


# ─────────────────────────────────────────────
# PATCH J: ROTATING FILE LOGGER
# ─────────────────────────────────────────────
# Design note: every print() in this file also needs to land in
# paper_trader.log so scheduled runs have an audit trail. Rather than
# converting hundreds of print() calls to logger.info(), we tee sys.stdout
# to a file with per-line timestamps. Preserves interactive console
# behavior, survives encoding issues, rotates on size.

LOG_FILE = Path(__file__).parent / "paper_trader.log"
_MAX_LOG_BYTES = 10 * 1024 * 1024   # 10MB before rotate
_LOG_BACKUPS = 5                     # keep 5 old files


def _rotate_log_if_needed():
    """Size-based rotation. Shifts .log → .log.1 → .log.2 → … and drops the
    oldest. Silent no-op if file doesn't exist or stat fails."""
    if not LOG_FILE.exists():
        return
    try:
        if LOG_FILE.stat().st_size < _MAX_LOG_BYTES:
            return
    except OSError:
        return
    for i in range(_LOG_BACKUPS - 1, 0, -1):
        src = LOG_FILE.with_suffix(f".log.{i}")
        dst = LOG_FILE.with_suffix(f".log.{i+1}")
        if src.exists():
            try:
                src.replace(dst)
            except OSError:
                return
    try:
        LOG_FILE.replace(LOG_FILE.with_suffix(".log.1"))
    except OSError:
        pass


class _TeeStdout:
    """Duplicate every stdout write to paper_trader.log with per-line
    timestamps. Never raises on write failure — a trading loop must not
    die because the disk is full."""

    def __init__(self, original):
        self._original = original
        self._buf = ""
        _rotate_log_if_needed()
        try:
            self._fh = open(LOG_FILE, "a", encoding="utf-8")
        except OSError:
            self._fh = None

    def write(self, data):
        self._original.write(data)
        if self._fh is None:
            return
        self._buf += data
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                self._fh.write(f"{ts}  {line}\n")
                self._fh.flush()
            except OSError:
                pass

    def flush(self):
        try:
            self._original.flush()
        except OSError:
            pass
        if self._fh is not None:
            try:
                self._fh.flush()
            except OSError:
                pass

    def isatty(self):
        return getattr(self._original, "isatty", lambda: False)()


def setup_file_logging():
    """Install the tee'd stdout. Idempotent — safe to call multiple times."""
    if isinstance(sys.stdout, _TeeStdout):
        return
    sys.stdout = _TeeStdout(sys.stdout)


# ─────────────────────────────────────────────
# PATCH I: CONFIG VALIDATION
# ─────────────────────────────────────────────
def validate_config():
    """Sanity-check CONFIG at startup. Aborts with a readable error on any
    violation so a bad edit doesn't silently change trading behavior.

    Caught issues include: swapped min/max prices, fractions outside [0,1],
    drawdown_soft ≥ drawdown_hard (would disable soft throttling), negative
    bankrolls, etc.
    """
    errors = []

    def want_range(key, lo, hi, inclusive=True):
        v = CONFIG.get(key)
        if v is None:
            errors.append(f"{key} missing from CONFIG")
            return
        if inclusive:
            if v < lo or v > hi:
                errors.append(f"{key}={v} out of range [{lo}, {hi}]")
        else:
            if v <= lo or v >= hi:
                errors.append(f"{key}={v} out of range ({lo}, {hi})")

    def want_positive(key):
        v = CONFIG.get(key)
        if v is None or v <= 0:
            errors.append(f"{key}={v} must be > 0")

    # Price + EV bounds
    for k in ["min_ev", "min_price", "max_price", "min_edge_narrow_bin",
              "min_model_prob", "kelly_fraction", "max_bet_frac",
              "max_daily_exposure", "drawdown_soft", "drawdown_hard",
              "prob_shrinkage"]:
        want_range(k, 0.0, 1.0)

    # Ordering checks
    if CONFIG.get("min_price", 0) >= CONFIG.get("max_price", 1):
        errors.append(
            f"min_price ({CONFIG.get('min_price')}) must be < max_price "
            f"({CONFIG.get('max_price')})"
        )
    if CONFIG.get("drawdown_soft", 0) >= CONFIG.get("drawdown_hard", 1):
        errors.append(
            f"drawdown_soft ({CONFIG.get('drawdown_soft')}) must be < "
            f"drawdown_hard ({CONFIG.get('drawdown_hard')})"
        )
    if CONFIG.get("min_hours", 0) >= CONFIG.get("max_hours", 1):
        errors.append(
            f"min_hours ({CONFIG.get('min_hours')}) must be < max_hours "
            f"({CONFIG.get('max_hours')})"
        )
    if CONFIG.get("min_ev", 0) > CONFIG.get("min_edge_narrow_bin", 1):
        errors.append(
            f"min_edge_narrow_bin ({CONFIG.get('min_edge_narrow_bin')}) "
            f"should be ≥ min_ev ({CONFIG.get('min_ev')}) — narrow bins "
            f"need MORE edge to trade, not less"
        )

    # Positive integers/floats
    for k in ["starting_bankroll", "max_shares", "min_volume",
              "min_calibration_obs"]:
        want_positive(k)

    if errors:
        msg = "CONFIG validation failed:\n  - " + "\n  - ".join(errors)
        raise ValueError(msg)


# ─────────────────────────────────────────────
# TRADES LOG (JSON source of truth)
# ─────────────────────────────────────────────
def load_trades():
    if TRADES_FILE.exists():
        try:
            with open(TRADES_FILE) as f:
                return json.load(f)
        except:
            return []
    return []


def save_trades(trades):
    # Atomic write: dump to a temp file then rename. Prevents mid-write
    # truncation (which we hit multiple times — a SIGTERM during json.dump
    # left paper_trades.json corrupt with a half-written final object).
    tmp = TRADES_FILE.with_suffix(TRADES_FILE.suffix + ".tmp")
    with open(tmp, "w") as f:
        json.dump(trades, f, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, TRADES_FILE)


# ─────────────────────────────────────────────
# BANKROLL STATE (PATCH F)
# ─────────────────────────────────────────────
def compute_current_bankroll(trades):
    """Live bankroll = starting + realized P&L. Open trades are NOT marked
    to market (we don't have real-time mid-prices). Only resolved trades
    move bankroll. Voided trades have pnl=0 so they naturally contribute
    nothing — still counted as 'resolved' for exposure purposes."""
    resolved = [t for t in trades if t.get("resolved")]
    pnl = sum(t.get("pnl", 0) for t in resolved)
    return CONFIG["starting_bankroll"] + pnl


def compute_open_exposure(trades):
    """Total $ currently at risk across unresolved trades. PATCH K:
    voided trades (market never closed) are 'resolved' and therefore
    excluded here — their cost is released back to the daily budget."""
    return sum(t.get("cost", 0) for t in trades if not t.get("resolved"))


def drawdown_factor(bankroll):
    """Return a 0.0-1.0 multiplier on Kelly based on drawdown from start.
       - < soft threshold (20%): full Kelly
       - between soft and hard: 0.5 × Kelly
       - ≥ hard threshold (40%): 0.0 (stop trading)
    """
    start = CONFIG["starting_bankroll"]
    if bankroll >= start:
        return 1.0
    dd = (start - bankroll) / start
    if dd >= CONFIG["drawdown_hard"]:
        return 0.0
    if dd >= CONFIG["drawdown_soft"]:
        return 0.5
    return 1.0


def prob_shrink(model_prob, market_prob):
    """Pull model prob toward market by `prob_shrinkage`. Margin of safety:
    if our σ is slightly off, the implied edge is slightly off too — shrinking
    toward the market price discounts the bet size proportionally."""
    s = CONFIG["prob_shrinkage"]
    return model_prob * (1 - s) + market_prob * s


# ─────────────────────────────────────────────
# PATCH G: PER-CITY KELLY SCALING FROM BRIER SCORES
# ─────────────────────────────────────────────
# Not every city forecasts equally well. Our Brier backtest (run via
# brier_backtest.py on 85-day paired forecast/actual windows) produces a
# Brier Skill Score per city, which tells us how much better our model
# is than climatology. Cities with BSS near 0 have no real edge to extract;
# cities with BSS > 0.25 are genuinely predictive. PATCH G multiplies the
# Kelly scalar per city so capital flows toward where our edge actually lives.

BSS_PATH = Path(__file__).parent / "backtest_metar.csv"

def load_city_bss():
    """Read per-city Brier Skill Score from the most recent backtest CSV.
    Returns a dict: {slug: bss_float}. Missing file → empty dict (safe — we
    then fall back to a cautious 0.5 scalar for every city)."""
    bss_map = {}
    if not BSS_PATH.exists():
        return bss_map
    try:
        with open(BSS_PATH) as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    bss_map[row["slug"]] = float(row["bss"])
                except (ValueError, KeyError, TypeError):
                    pass
    except Exception:
        return {}
    return bss_map


CITY_BSS = load_city_bss()


def kelly_scalar_for_city(slug):
    """Scale Kelly per city based on that city's Brier Skill Score.

    Bands were chosen after eyeballing the BSS distribution — there's a
    real cluster at BSS ≥ 0.35 (the top-tier cities) and a second cluster
    around 0.15-0.25 (the solid-but-noisier set), with Singapore alone
    below 0.05.

      BSS ≥ 0.25   → 1.00×  (full trust — NYC, Chicago, Warsaw, etc.)
      BSS 0.15-0.25 → 0.75× (good but less certain — SF, Miami, Seoul)
      BSS 0.05-0.15 → 0.50× (marginal edge — Hong Kong)
      BSS < 0.05    → 0.00× (no real skill — block trading — Singapore)
      missing slug  → 0.50× (cautious default for cities not yet backtested)
    """
    if slug not in CITY_BSS:
        return 0.5
    bss = CITY_BSS[slug]
    if bss >= 0.25:
        return 1.0
    if bss >= 0.15:
        return 0.75
    if bss >= 0.05:
        return 0.50
    return 0.0


# ─────────────────────────────────────────────
# PROBABILITY MODEL (copied from app.py core)
# ─────────────────────────────────────────────
def parse_bin_range(label, unit="F"):
    label = label.strip()
    if "or higher" in label.lower() or "or above" in label.lower():
        nums = re.findall(r'-?\d+', label)
        return (int(nums[0]), int(nums[0]) + 30) if nums else (None, None)
    if "or below" in label.lower() or "or lower" in label.lower():
        nums = re.findall(r'-?\d+', label)
        return (int(nums[0]) - 30, int(nums[0])) if nums else (None, None)
    range_match = re.match(r'(-?\d+)\s*(?:to|–|—)\s*(-?\d+)', label)
    if range_match:
        return (int(range_match.group(1)), int(range_match.group(2)))
    hyphen_match = re.match(r'(-?\d+)\s*-\s*(-?\d+)', label)
    if hyphen_match:
        lo, hi = int(hyphen_match.group(1)), int(hyphen_match.group(2))
        if hi >= lo:
            return (lo, hi)
        return (min(lo, hi), max(lo, hi))
    nums = re.findall(r'-?\d+', label)
    if len(nums) >= 2:
        return (min(int(nums[0]), int(nums[1])), max(int(nums[0]), int(nums[1])))
    elif len(nums) == 1:
        return (int(nums[0]), int(nums[0]))
    return (None, None)


def build_model_probabilities(forecast_high, bins, unit="F",
                               model_spread=None, empirical_std=None,
                               historical=None, lead_days=0):
    """
    Build per-bin probabilities for a daily-high market.

    PATCH C: the prior version blended climatological std (historical spread of
    same-calendar-day highs across years) at 0.4 weight and added a 0.25 KDE-
    over-historical weight on top. That mixes unconditional seasonal variance
    into what should be a forecast-conditional distribution, which systemati-
    cally fattens the tails. On days where HRRR/ECMWF/GFS agree within ~2°F,
    σ should be ~1-2°F, not ~5°F. Now we use forecast-conditional σ only.

    PATCH D: σ is scaled by lead time. D+0 / D+1 / D+2 forecast RMSE is
    roughly 1.0 / 1.5 / 2.0°F in well-behaved regimes.

    PATCH H: the 0.3-per-lead scaling from PATCH D turned out to be too
    tight. Resolved-trade analysis (24 trades, Apr 2026) showed D+1
    predicted 34.5% win rate but actual was 7.7% — a 27-point calibration
    gap driven by systematically overconfident σ at longer leads.
    Bumping the scaling factor to 0.5-per-lead puts D+1 σ at 1.5× D+0 and
    D+2 at 2.0×, which lines up with operational forecast RMSE literature
    (σ roughly doubles from D+0 to D+2). This reduces model_prob on deep
    bins at longer leads and blocks the "edge" that wasn't real.

    Historical data is still used — but only for skew-shape, and only when
    we have enough of it (n ≥ 20) to make skew estimation stable.
    """
    # 1) Pick base_std from forecast-conditional sources only.
    if empirical_std is not None and empirical_std > 0:
        base_std = empirical_std
    elif model_spread is not None and model_spread > 1.0:
        base_std = model_spread * 1.5
    else:
        base_std = 1.5 if unit == "F" else 0.8  # D+0 floor

    # 2) Scale with lead time (more uncertainty further out). PATCH H: 0.5
    #    per lead, not 0.3, after live data showed D+1 σ was too narrow.
    effective_std = base_std * (1.0 + 0.5 * max(0, lead_days))

    # 3) Hard floor so we never imply super-confident forecasts.
    effective_std = max(effective_std, 1.0 if unit == "F" else 0.6)

    # 4) Skew only if we have a stable sample.
    skew_param = 0
    if historical and len(historical) >= 20:
        hist_vals = np.array([h["high_f"] if unit == "F" else h["high_c"] for h in historical])
        skew_param = np.clip(stats.skew(hist_vals), -2, 2)

    dist = skewnorm(a=skew_param, loc=forecast_high, scale=effective_std)

    probs = {}
    for b in bins:
        lo, hi = parse_bin_range(b, unit)
        if lo is not None and hi is not None:
            probs[b] = dist.cdf(hi + 0.5) - dist.cdf(lo - 0.5)

    total = sum(probs.values())
    if total > 0:
        probs = {k: v / total for k, v in probs.items()}
    return probs


def get_forecast_from_log(log, target_str, max_age_hours=4.0, min_models=3):
    """Return (forecast, model_highs_dict) for target_str from an already-loaded
    forecast log, or (None, None) if missing / incomplete / stale.

    forecast_logger.py writes these fields every cycle (runs seconds before
    paper_trader in the pipeline), so in the common case we can skip the
    NWS + Open-Meteo fetches entirely — roughly 6-7 API calls saved per
    (city, day) pair, which is the hot path.

    Safety rails (both cause the caller to fall back to the live API path):
      - max_age_hours: reject entries older than this. Weather models refresh
        every 6 hours, so a stale log can be meaningfully wrong near fronts.
        4h gives one scheduler cycle of headroom before we treat data as old.
      - min_models: require at least this many ensemble members in model_highs.
        model_spread is computed as np.std(model_highs.values()) and used as
        the σ input to the probability model — with too few samples it
        underestimates uncertainty, inflating edge and manufacturing signal.

    Missing or unparseable updated_at → treated as stale (old log format)."""
    for entry in log:
        if entry.get("date") != target_str:
            continue
        forecast = entry.get("forecast")
        model_highs = entry.get("model_highs") or {}
        if forecast is None or not model_highs:
            return None, None
        # Gate 1: need enough ensemble members for a meaningful spread.
        if len(model_highs) < min_models:
            return None, None
        # Gate 2: freshness. Weather forecasts decay; don't reuse stale entries.
        updated_at = entry.get("updated_at")
        if not updated_at:
            return None, None
        try:
            ts = datetime.fromisoformat(updated_at)
        except (ValueError, TypeError):
            return None, None
        age_hours = (datetime.now() - ts).total_seconds() / 3600.0
        if age_hours > max_age_hours:
            return None, None
        # Return a copy so caller mutations don't leak back into the log.
        return forecast, dict(model_highs)
    return None, None


def get_historical_highs(lat, lon, target_date_str, unit="F"):
    month_day = target_date_str[5:]
    records = []
    for year in range(2015, 2026):
        target = f"{year}-{month_day}"
        start = (datetime.strptime(target, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        end = (datetime.strptime(target, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        try:
            r = retry_get(
                f"https://archive-api.open-meteo.com/v1/archive?latitude={lat}&longitude={lon}"
                f"&daily=temperature_2m_max&temperature_unit=fahrenheit&timezone=auto"
                f"&start_date={start}&end_date={end}", timeout=8)
            if r.status_code == 200:
                d = r.json()
                if "daily" in d and d["daily"]["temperature_2m_max"]:
                    idx = 1 if len(d["daily"]["temperature_2m_max"]) > 1 else 0
                    high = d["daily"]["temperature_2m_max"][idx]
                    if high is not None:
                        records.append({
                            "year": year,
                            "high_f": round(high),
                            "high_c": round((high - 32) * 5 / 9)
                        })
        except:
            pass
    return records


# ─────────────────────────────────────────────
# MARKET PARSING
# ─────────────────────────────────────────────
def parse_event_markets(event):
    if not event:
        return []
    markets = []
    for m in event.get("markets", []):
        title = m.get("groupItemTitle", "")
        prices_raw = m.get("outcomePrices", "[]")
        if isinstance(prices_raw, str):
            try:
                prices = json.loads(prices_raw)
            except:
                prices = []
        else:
            prices = prices_raw
        yes_price = float(prices[0]) if prices else 0
        resolved_yes = (str(prices[0]) == "1") if prices else False
        markets.append({
            "bin": title,
            "yes_price": yes_price,
            "volume": m.get("volumeNum", 0),
            "closed": m.get("closed", False),
            "resolved_yes": resolved_yes,
        })
    return sorted(markets, key=lambda x: parse_bin_range(x["bin"])[0] or 0)


# ─────────────────────────────────────────────
# EDGE SCANNER
# ─────────────────────────────────────────────
def scan_city_edges(city_name, city, verbose=True,
                    bankroll=None, open_exposure=0.0):
    """Scan active markets for a city and return trade signals.

    PATCH F: bankroll + open_exposure are threaded in from main() so that
    sizing scales with live capital and respects a portfolio-level
    exposure cap. When bankroll is None we fall back to starting_bankroll
    (useful for unit-testing the scanner in isolation)."""
    slug = city["slug"]
    today = city_today(city)
    signals = []

    if bankroll is None:
        bankroll = CONFIG["starting_bankroll"]

    # Drawdown circuit breaker: zero factor → don't even scan.
    dd_factor = drawdown_factor(bankroll)
    if dd_factor == 0.0:
        if verbose:
            print(f"  🛑 [{slug}] skipped — bankroll ${bankroll:.2f} "
                  f"below hard drawdown floor "
                  f"(${CONFIG['starting_bankroll'] * (1 - CONFIG['drawdown_hard']):.2f})")
        return []

    # PATCH G: per-city Kelly scaling. If BSS is too low, the model has no
    # real edge here — don't scan at all. Saves API calls and prevents the
    # model from "finding" phantom edges in noise.
    city_scalar = kelly_scalar_for_city(slug)
    if city_scalar == 0.0:
        if verbose:
            bss = CITY_BSS.get(slug)
            bss_str = f"{bss:.3f}" if bss is not None else "unscored"
            print(f"  ⏸  [{slug}] blocked — BSS={bss_str} below floor (no model skill)")
        return []

    # Per-trade cap and remaining daily exposure budget are both in $.
    per_trade_cap = CONFIG["max_bet_frac"] * bankroll
    daily_budget = CONFIG["max_daily_exposure"] * bankroll
    remaining_budget = max(0.0, daily_budget - open_exposure)

    # Load calibration data
    log = load_forecast_log(slug)

    # PATCH B: sanity-filter paired observations before computing bias/std.
    # A single corrupt entry (e.g. actual=-0.5 from a bin-midpoint parse bug)
    # can poison the mean and blow up the std, which in turn inflates
    # effective_std in the probability model and manufactures fake "edge"
    # on deep OTM tail bins. Reject any pair where the actual is outside
    # physically plausible bounds OR where |actual - forecast| exceeds a
    # threshold consistent with a data error rather than a forecast miss.
    max_plausible_error = 10.0 if city["unit"] == "F" else 6.0  # °F / °C
    errors = [
        e["actual"] - e["forecast"]
        for e in log
        if e.get("forecast") is not None
        and e.get("actual") is not None
        and is_sane_temp(e["actual"], city["unit"])
        and abs(e["actual"] - e["forecast"]) <= max_plausible_error
    ]
    n_obs = len(errors)
    empirical_bias = round(np.mean(errors), 2) if n_obs >= 3 else None
    empirical_std = round(np.std(errors), 2) if n_obs >= 3 else None

    # GATE 1: Block cities without enough calibration data.
    # Trading uncalibrated cities is how we lost the Taipei bets — we had
    # no idea if our 5% "edge" was real or a model artifact.
    wu_offset = city.get("wu_offset", 0)
    has_station_cal = abs(wu_offset) > 0.01  # City has had station offset measured
    if n_obs < CONFIG["min_calibration_obs"] and not has_station_cal:
        if verbose:
            print(f"  ⏸  [{slug}] skipped — only {n_obs} paired observations "
                  f"(need {CONFIG['min_calibration_obs']}) and no station offset")
        return []

    for days_ahead in range(0, 3):
        target_date = today + timedelta(days=days_ahead)
        target_str = target_date.strftime("%Y-%m-%d")

        event_slug = build_event_slug(slug, target_date)
        event = fetch_polymarket_event(event_slug)
        if event is None or event.get("closed"):
            continue

        markets = parse_event_markets(event)
        if not markets:
            continue

        # Check volume
        total_vol = sum(m["volume"] for m in markets)
        if total_vol < CONFIG["min_volume"]:
            continue

        # Fast path: reuse the ensemble + model_highs that forecast_logger.py
        # wrote to forecast_logs/{slug}.json seconds ago. Saves ~7 API calls per
        # (city, day) pair. model_highs in the log already includes NWS under
        # the "nws" key for F-unit cities, so we don't need to re-fetch.
        ensemble_high, converted_highs = get_forecast_from_log(log, target_str)

        # Fallback: if the log is missing/stale for this date (forecast_logger
        # didn't run, or errored out for this city), hit the APIs the old way.
        # NWS removed — analysis showed it had 3x worse RMSE than GFS everywhere.
        if ensemble_high is None:
            model_highs = fetch_openmeteo_multimodel_highs(
                city["lat"], city["lon"], target_str)
            ensemble_high, converted_highs = compute_ensemble_high(
                model_highs, days_ahead, city["unit"], slug=slug
            )

        if ensemble_high is None:
            continue

        # Apply WU station offset (systematic difference between our data source and Polymarket resolution)
        if wu_offset != 0:
            ensemble_high = round(ensemble_high + wu_offset, 1)

        # Apply empirical bias correction (forecast vs actual from calibration loop)
        if empirical_bias is not None:
            ensemble_high = round(ensemble_high + empirical_bias, 1)

        # Model spread for uncertainty
        spread_vals = list(converted_highs.values())
        model_spread = np.std(spread_vals) if len(spread_vals) >= 2 else None

        # Historical data for distribution shape
        historical = get_historical_highs(city["lat"], city["lon"], target_str, city["unit"])

        # Build probability model (PATCH D: pass lead_days so σ can scale)
        bin_labels = [m["bin"] for m in markets]
        model_probs = build_model_probabilities(
            ensemble_high, bin_labels, city["unit"],
            model_spread=model_spread, empirical_std=empirical_std,
            historical=historical, lead_days=days_ahead,
        )

        # Find edges
        for m in markets:
            b = m["bin"]
            mp = model_probs.get(b, 0)
            mkt = m["yes_price"]

            # GATE 2: Price range. Min 8¢ blocks long-shot lottery tickets.
            # Real trades showed: 0.5¢-6¢ bins LOST consistently. Our model
            # is not accurate enough to price sub-8¢ events correctly.
            if mkt < CONFIG["min_price"] or mkt >= CONFIG["max_price"]:
                continue

            # GATE 2b (PATCH E): mirror the price floor on the model side.
            # If we think the bin has <3% chance, refuse to trade it regardless
            # of "edge" — that's where σ misestimates show up as fake signal.
            if mp < CONFIG["min_model_prob"]:
                continue

            # GATE 3: Narrow bins (2°F or 1°C) are structurally hard to hit.
            # Require a much bigger edge (12%) to bet them.
            lo, hi = parse_bin_range(b, city["unit"])
            is_narrow_bin = (lo is not None and hi is not None
                             and (hi - lo) <= (2 if city["unit"] == "F" else 1))

            # PATCH F: shrink model prob toward market before computing edge/Kelly.
            # This gives a margin of safety — we deliberately size off a slightly
            # discounted edge to absorb σ estimation error.
            mp_sized = prob_shrink(mp, mkt)

            edge = mp - mkt  # reported edge stays raw (what the model actually says)
            edge_sized = mp_sized - mkt  # edge used for sizing decisions
            min_edge = CONFIG["min_edge_narrow_bin"] if is_narrow_bin else CONFIG["min_ev"]
            if edge < min_edge:
                continue

            # If shrinkage pulls the edge below zero, don't bet.
            if edge_sized <= 0:
                continue

            # Kelly sizing on the shrunk probability, scaled by (a) drawdown
            # factor and (b) per-city Kelly scalar from BSS. Both multipliers
            # are ≤ 1.0, so bet sizes only ever shrink from raw Kelly — never
            # grow.
            b_odds = (1 - mkt) / mkt
            kelly = max(0, (mp_sized * b_odds - (1 - mp_sized)) / b_odds)
            quarter_kelly = kelly * CONFIG["kelly_fraction"] * dd_factor * city_scalar

            # Sizing caps (all fractions of LIVE bankroll, not a static constant):
            #   - per_trade_cap  : single-trade max ($ from max_bet_frac × bankroll)
            #   - remaining_budget: daily exposure budget minus already-open trades
            raw_bet = quarter_kelly * bankroll
            bet_size = min(raw_bet, per_trade_cap, remaining_budget)

            if bet_size < 1.00:  # Too small to bother with — even if budget-limited
                continue

            shares = round(bet_size / mkt)

            # GATE 4: Cap share count. 4000-share bets on 0.5¢ tickets are insane.
            if shares > CONFIG["max_shares"]:
                shares = CONFIG["max_shares"]

            cost = round(shares * mkt, 2)

            # After share-rounding, the actual cost can still overshoot the
            # remaining budget by a few cents. Drop anything that breaches it.
            if cost > remaining_budget + 0.01:
                continue

            # Consume daily budget so downstream signals in this scan see
            # the running total. (open_exposure tracks pre-existing trades;
            # remaining_budget tracks what's left to commit right now.)
            remaining_budget -= cost

            signal = {
                "city": city_name,
                "slug": slug,
                "station": city["station"],
                "unit": city["unit"],
                "market_date": target_str,
                "lead_days": days_ahead,
                "bin": b,
                "entry_price": round(mkt * 100, 1),
                "shares": shares,
                "cost": cost,
                "forecast_high": ensemble_high,
                "model_prob": round(mp * 100, 1),
                "model_prob_shrunk": round(mp_sized * 100, 1),  # PATCH F
                "market_prob": round(mkt * 100, 1),
                "edge_pct": round(edge * 100, 1),
                "edge_sized_pct": round(edge_sized * 100, 1),   # PATCH F
                "kelly_pct": round(quarter_kelly * 100, 1),
                "dd_factor": dd_factor,                          # PATCH F
                "bankroll_at_entry": round(bankroll, 2),         # PATCH F
                "city_kelly_scalar": city_scalar,                # PATCH G
                "city_bss": CITY_BSS.get(slug),                  # PATCH G
                "n_models": len(converted_highs),
                "model_spread": round(model_spread, 1) if model_spread else None,
                "calibrated": empirical_std is not None,
            }
            signals.append(signal)

            if verbose:
                cal_tag = "CAL" if empirical_std is not None else "---"
                print(f"  🎯 [{slug}] {target_str} D+{days_ahead}: "
                      f"{b} @ {signal['entry_price']}¢ | "
                      f"edge={signal['edge_pct']:+.1f}% kelly={signal['kelly_pct']:.1f}% "
                      f"cost=${cost:.2f} [{cal_tag}]")

        time.sleep(0.3)

    return signals


STALE_TRADE_DAYS = 5  # PATCH K: mark as voided if market_date is this many days past


def resolve_trades(trades, verbose=True):
    """Check resolved markets and fill in outcomes for open trades.

    PATCH K: also voids stale trades — if market_date is more than
    STALE_TRADE_DAYS days in the past and still unresolved, we assume the
    Polymarket market was cancelled or silently expired. Mark it
    voided=True with zero P&L so the exposure frees up for new signals.
    Without this, a single cancelled market can park $15-50 of our daily
    budget forever."""
    resolved_count = 0
    voided_count = 0
    today = datetime.now().date()

    for trade in trades:
        if trade.get("resolved"):
            continue

        slug = trade["slug"]
        market_date = trade["market_date"]
        d = datetime.strptime(market_date, "%Y-%m-%d").date()

        event_slug = build_event_slug(slug, d)
        event = fetch_polymarket_event(event_slug)

        if event is None or not event.get("closed"):
            # PATCH K: if we can't find a closed market but it's been days,
            # void the trade so exposure doesn't stay stuck forever.
            days_past = (today - d).days
            if days_past >= STALE_TRADE_DAYS:
                trade["outcome_bin"] = None
                trade["resolved_price"] = None
                trade["revenue"] = 0.0
                trade["pnl"] = 0.0  # break-even — don't count as win or loss
                trade["return_pct"] = 0.0
                trade["resolved"] = True
                trade["voided"] = True
                trade["resolved_at"] = datetime.now().isoformat()
                voided_count += 1
                if verbose:
                    print(f"  🗑  [{slug}] {market_date}: {trade['bin']} "
                          f"→ VOIDED ({days_past}d past, market never closed). "
                          f"Cost ${trade['cost']:.2f} freed from exposure.")
            continue

        winner = get_resolution_winner(event)
        if winner is None:
            continue

        won = (winner == trade["bin"])
        resolved_price = 100 if won else 0
        revenue = round(resolved_price / 100 * trade["shares"], 2)
        pnl = round(revenue - trade["cost"], 2)
        ret = round(pnl / trade["cost"] * 100, 1) if trade["cost"] > 0 else 0

        trade["outcome_bin"] = winner
        trade["resolved_price"] = resolved_price
        trade["revenue"] = revenue
        trade["pnl"] = pnl
        trade["return_pct"] = ret
        trade["resolved"] = True
        trade["resolved_at"] = datetime.now().isoformat()

        resolved_count += 1
        emoji = "✅" if won else "❌"
        if verbose:
            print(f"  {emoji} [{slug}] {market_date}: {trade['bin']} → "
                  f"resolved={winner} | P&L=${pnl:+.2f} ({ret:+.1f}%)")

        time.sleep(0.2)

    if voided_count and verbose:
        print(f"  (voided {voided_count} stale trade(s) to free exposure budget)")

    return resolved_count


# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────
def export_to_excel(trades):
    """Write trades to the paper_trade_tracker.xlsx, preserving formatting."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  openpyxl not installed — skipping Excel export. pip install openpyxl")
        return

    if not EXCEL_FILE.exists():
        print(f"  Excel template not found at {EXCEL_FILE} — skipping export")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Trade Log"]

    for i, trade in enumerate(trades[:50]):
        row = 6 + i
        ws.cell(row=row, column=2).value = trade.get("logged_at", "")[:10]  # Date
        ws.cell(row=row, column=3).value = trade["city"]                     # City
        # D: Station auto-fills via VLOOKUP
        ws.cell(row=row, column=5).value = trade["market_date"]              # Market Date
        ws.cell(row=row, column=6).value = trade["bin"]                      # Bin
        ws.cell(row=row, column=7).value = trade["entry_price"]              # Entry Price ¢
        ws.cell(row=row, column=8).value = trade["shares"]                   # Shares
        # I: Cost auto-calculates
        ws.cell(row=row, column=10).value = trade["forecast_high"]           # Forecast High
        ws.cell(row=row, column=11).value = trade["model_prob"] / 100        # Model Prob
        ws.cell(row=row, column=12).value = trade["market_prob"] / 100       # Market Prob
        # M, N: Edge and Kelly auto-calculate

        if trade.get("resolved"):
            ws.cell(row=row, column=15).value = trade.get("outcome_bin", "")    # Outcome Bin
            ws.cell(row=row, column=16).value = trade.get("resolved_price", "") # Resolved Price
            # Q, R, S: Revenue, P&L, Return auto-calculate

        notes = []
        if trade.get("calibrated"):
            notes.append("calibrated")
        notes.append(f"D+{trade['lead_days']}")
        notes.append(f"{trade['n_models']}mod")
        if trade.get("model_spread"):
            notes.append(f"spread={trade['model_spread']}°")
        ws.cell(row=row, column=20).value = " | ".join(notes)

    try:
        wb.save(EXCEL_FILE)
        print(f"  Exported {len(trades[:50])} trades to {EXCEL_FILE.name}")
    except PermissionError:
        # Excel is open in another window and holding a lock on the file.
        # The JSON source of truth was already saved; Excel is a pretty mirror,
        # so losing this write is cosmetic. Don't kill the run over it.
        print(f"  ⚠  {EXCEL_FILE.name} is open in Excel — skipping export. "
              f"Close Excel and re-run `python paper_trader.py --export` "
              f"to sync the spreadsheet.")
    except Exception as e:
        print(f"  ⚠  Excel export failed: {type(e).__name__}: {e}")


# ─────────────────────────────────────────────
# PERFORMANCE SUMMARY
# ─────────────────────────────────────────────
def print_performance():
    trades = load_trades()
    total = len(trades)
    # PATCH K: voided trades are "resolved" for exposure purposes but aren't
    # genuine outcomes — exclude them from win/loss stats so voided markets
    # don't pollute ROI.
    resolved_all = [t for t in trades if t.get("resolved")]
    voided = [t for t in resolved_all if t.get("voided")]
    resolved = [t for t in resolved_all if not t.get("voided")]
    open_trades = [t for t in trades if not t.get("resolved")]
    wins = [t for t in resolved if t.get("pnl", 0) > 0]
    losses = [t for t in resolved if t.get("pnl", 0) < 0]

    total_cost = sum(t["cost"] for t in resolved) if resolved else 0
    total_pnl = sum(t.get("pnl", 0) for t in resolved)
    total_revenue = sum(t.get("revenue", 0) for t in resolved)
    win_pnl = sum(t.get("pnl", 0) for t in wins) if wins else 0
    loss_pnl = abs(sum(t.get("pnl", 0) for t in losses)) if losses else 0

    print(f"\n{'='*60}")
    print(f"PAPER TRADING PERFORMANCE")
    print(f"{'='*60}")
    print(f"  Total signals logged:  {total}")
    print(f"  Open (unresolved):     {len(open_trades)}")
    print(f"  Resolved:              {len(resolved)}")
    if voided:
        print(f"  Voided (stale):        {len(voided)}")
    print(f"  Wins:                  {len(wins)}")
    print(f"  Losses:                {len(losses)}")
    print(f"  Win rate:              {len(wins)/len(resolved)*100:.1f}%" if resolved else "  Win rate:              ---")
    print(f"  Total cost:            ${total_cost:.2f}")
    print(f"  Total revenue:         ${total_revenue:.2f}")
    print(f"  Total P&L:             ${total_pnl:+.2f}")
    print(f"  ROI:                   {total_pnl/total_cost*100:+.1f}%" if total_cost > 0 else "  ROI:                   ---")
    print(f"  Avg win:               ${win_pnl/len(wins):.2f}" if wins else "  Avg win:               ---")
    print(f"  Avg loss:              ${loss_pnl/len(losses):.2f}" if losses else "  Avg loss:              ---")
    print(f"  Profit factor:         {win_pnl/loss_pnl:.2f}x" if loss_pnl > 0 else "  Profit factor:         ∞" if win_pnl > 0 else "  Profit factor:         ---")
    print(f"{'='*60}")

    if resolved:
        print(f"\n  Last 5 resolved trades:")
        for t in sorted(resolved, key=lambda x: x.get("resolved_at", ""), reverse=True)[:5]:
            emoji = "✅" if t.get("pnl", 0) > 0 else "❌"
            print(f"    {emoji} {t['city']} {t['market_date']} | {t['bin']} @ {t['entry_price']}¢ "
                  f"→ ${t.get('pnl', 0):+.2f} ({t.get('return_pct', 0):+.1f}%)")
    print()


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    import argparse
    parser = argparse.ArgumentParser(description="Polymarket Weather Auto Paper Trader")
    parser.add_argument("--status", action="store_true", help="Print performance summary")
    parser.add_argument("--export", action="store_true", help="Re-export JSON → Excel")
    parser.add_argument("--city", type=str, help="Scan a single city slug")
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args()

    # PATCH J: tee stdout to paper_trader.log (rotating, 10MB × 5).
    # Do this BEFORE any print() so the first line of every run is captured.
    setup_file_logging()

    # PATCH I: fail fast on bad CONFIG before doing any work.
    try:
        validate_config()
    except ValueError as e:
        print(f"❌ {e}")
        sys.exit(1)

    verbose = not args.quiet

    if args.status:
        print_performance()
        return

    trades = load_trades()

    if args.export:
        export_to_excel(trades)
        print_performance()
        return

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'='*60}")
    print(f"Paper Trader — {timestamp}")
    print(f"Config: min_ev={CONFIG['min_ev']*100:.0f}% max_price={CONFIG['max_price']*100:.0f}¢ "
          f"kelly={CONFIG['kelly_fraction']:.0%}")
    print(f"{'='*60}")

    # Step 1: Resolve any open trades
    open_trades = [t for t in trades if not t.get("resolved")]
    if open_trades:
        print(f"\n📋 Resolving {len(open_trades)} open trades...")
        n_resolved = resolve_trades(trades, verbose)
        save_trades(trades)
        if n_resolved > 0:
            print(f"  Resolved {n_resolved} trades")
    else:
        print(f"\n📋 No open trades to resolve")

    # PATCH F: Compute live bankroll state ONCE, AFTER resolution.
    # Resolution can move P&L, so sizing uses post-resolve bankroll.
    bankroll = compute_current_bankroll(trades)
    open_exposure = compute_open_exposure(trades)
    dd_factor = drawdown_factor(bankroll)
    per_trade_cap = CONFIG["max_bet_frac"] * bankroll
    daily_budget = CONFIG["max_daily_exposure"] * bankroll
    remaining_budget = max(0.0, daily_budget - open_exposure)

    print(f"\n💰 Bankroll state:")
    print(f"  Live bankroll:         ${bankroll:.2f}  "
          f"(start ${CONFIG['starting_bankroll']:.2f}, "
          f"realized P&L ${bankroll - CONFIG['starting_bankroll']:+.2f})")
    print(f"  Open exposure:         ${open_exposure:.2f}")
    print(f"  Per-trade cap:         ${per_trade_cap:.2f}  "
          f"({CONFIG['max_bet_frac']*100:.0f}% of live)")
    print(f"  Daily budget:          ${daily_budget:.2f}  "
          f"({CONFIG['max_daily_exposure']*100:.0f}% of live)")
    print(f"  Remaining budget:      ${remaining_budget:.2f}")
    dd_label = "OK" if dd_factor == 1.0 else ("HALVED" if dd_factor == 0.5 else "HALTED")
    print(f"  Drawdown factor:       {dd_factor}  ({dd_label})")

    # PATCH G: summarize per-city Kelly scaling so you can see what's sized
    # at full confidence vs. de-weighted vs. blocked.
    if CITY_BSS:
        buckets = {"1.00×": [], "0.75×": [], "0.50×": [], "BLOCKED": []}
        unscored = []
        for name, c in CITIES.items():
            s = c["slug"]
            if s not in CITY_BSS:
                unscored.append(s)
                continue
            k = kelly_scalar_for_city(s)
            if   k == 1.0:  buckets["1.00×"].append(s)
            elif k == 0.75: buckets["0.75×"].append(s)
            elif k == 0.5:  buckets["0.50×"].append(s)
            else:           buckets["BLOCKED"].append(s)
        print(f"\n🎚  Per-city Kelly scaling (PATCH G):")
        for label, cities_in_bucket in buckets.items():
            if cities_in_bucket:
                print(f"  {label:8s}  ({len(cities_in_bucket):2d}): "
                      f"{', '.join(cities_in_bucket)}")
        if unscored:
            print(f"  UNSCORED ({len(unscored):2d}): {', '.join(unscored)}  "
                  f"→ default 0.50× (rerun brier_backtest.py to score)")
    else:
        print(f"\n⚠  PATCH G: no BSS data loaded from {BSS_PATH.name} — "
              f"all cities default to 0.50× Kelly")

    # Step 2: Scan for new signals
    if args.city:
        cities_to_scan = {n: c for n, c in CITIES.items() if c["slug"] == args.city}
    else:
        cities_to_scan = CITIES

    if dd_factor == 0.0:
        print(f"\n🛑 Drawdown hard stop hit. Skipping edge scan.")
        cities_to_scan = {}

    print(f"\n🔍 Scanning {len(cities_to_scan)} cities for edges...")

    # Track existing open trades to avoid duplicates
    existing_keys = {
        (t["slug"], t["market_date"], t["bin"])
        for t in trades if not t.get("resolved")
    }

    new_signals = 0
    running_exposure = open_exposure

    # Scan cities in parallel. Each city sees the SAME starting open_exposure —
    # they can't observe each other's commits during the scan. The reconciliation
    # loop below commits signals in deterministic city order and enforces the
    # daily budget cap signal-by-signal, mirroring serial behavior for the final
    # trade list (though individual trades may be sized slightly differently,
    # since all cities see the initial budget rather than a shrinking one).
    def _scan_one(city_name, city):
        try:
            signals = scan_city_edges(
                city_name, city, verbose,
                bankroll=bankroll,
                open_exposure=open_exposure,
            )
            return city_name, signals, None
        except Exception as e:
            return city_name, [], e

    results_by_city = {}  # city_name -> list of signals
    with ThreadPoolExecutor(max_workers=SCAN_WORKERS) as pool:
        futures = [pool.submit(_scan_one, name, city)
                   for name, city in cities_to_scan.items()]
        for fut in as_completed(futures):
            city_name, signals, err = fut.result()
            if err:
                with _scan_print_lock:
                    print(f"  [{city_name}] scan ERROR: {err!r}", flush=True)
                continue
            results_by_city[city_name] = signals

    # Commit signals in CITIES dict order, respecting the daily budget.
    for city_name in cities_to_scan.keys():
        if running_exposure >= daily_budget:
            if verbose:
                print(f"  🧯 Daily exposure budget exhausted at "
                      f"${running_exposure:.2f}/${daily_budget:.2f} — "
                      f"skipping remaining cities' signals")
            break
        for signal in results_by_city.get(city_name, []):
            key = (signal["slug"], signal["market_date"], signal["bin"])
            if key in existing_keys:
                continue
            # Signal-by-signal budget cap (cities all saw starting open_exposure,
            # so in aggregate they may have produced > daily_budget worth of signals)
            if running_exposure + signal["cost"] > daily_budget + 0.01:
                if verbose:
                    print(f"  🧯 [{signal['slug']}] {signal['bin']} dropped — "
                          f"would exceed daily budget "
                          f"(${running_exposure:.2f}+${signal['cost']:.2f} "
                          f"> ${daily_budget:.2f})")
                continue
            signal["logged_at"] = datetime.now().isoformat()
            signal["resolved"] = False
            trades.append(signal)
            existing_keys.add(key)
            running_exposure += signal["cost"]
            new_signals += 1

    save_trades(trades)

    print(f"\n{'='*60}")
    print(f"Done: {new_signals} new signals logged")
    print(f"{'='*60}")

    # Step 3: Export to Excel
    export_to_excel(trades)

    # Step 4: Performance summary
    print_performance()


if __name__ == "__main__":
    main()
